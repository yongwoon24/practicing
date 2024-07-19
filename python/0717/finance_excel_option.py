import pandas as pd
import time
from openpyxl import load_workbook
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options


options = Options()
user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
options.add_argument('user-agent=' + user_agent)
while True:
    select_theme = input('1:거래량 / 2:상승 / 3:하락 / 4:시가총액  >> ')
    if select_theme in ['1', '2', '3', '4']:
        break
    print('1~4 숫자를 입력하세요')

while True:
    select_kind = input('1:코스피 / 2:코스닥 >> ')
    if select_kind in ['1', '2']:
        break
    print('1 or 2 를 입력하세요')
select_theme_array = ['거래량', '상승', '하락', '시가총액']
driver = webdriver.Chrome(options=options)
driver.get("https://naver.com")
shortcut_link = driver.find_element(By.CSS_SELECTOR, '#shortcutArea > ul > li:nth-child(6) > a')
shortcut_link.click()

new_tab = driver.window_handles[-1]
driver.switch_to.window(new_tab)
driver.find_element(By.CSS_SELECTOR, f'#content > div.article > div.section > div.section_sise_top > ul > li.tab{select_theme} > a').click()
time.sleep(1)
menu_link = driver.find_element(By.CSS_SELECTOR, '#content > div.article > div.section > div.section_sise_top > div.group_type.is_active > a')
menu_link.click()
if(select_kind == '2'):
    driver.find_element(By.CSS_SELECTOR, '#contentarea > div.box_type_l > div > div:nth-child(2) > a').click()

rows = driver.find_elements(By.CSS_SELECTOR, 'table.type_2 tbody tr')

data = []

for row in rows:
    cells = row.find_elements(By.CSS_SELECTOR, 'td')
    if len(cells) >= 11:
        item = cells[1].text.strip()
        current_price = cells[2].text.strip()
        yday = cells[3].text.strip()
        change_rate = cells[4].text.strip()
        trade_amount = cells[5].text.strip()
        trade_price = cells[6].text.strip()
        buy_price = cells[7].text.strip()
        sell_price = cells[8].text.strip()
        total_price = cells[9].text.strip()
        per = cells[10].text.strip()
        roe = cells[11].text.strip()
        data.append([item, current_price, yday, change_rate, trade_amount, trade_price, buy_price, sell_price, total_price, per, roe])

driver.quit()

df = pd.DataFrame(data, columns=['종목명', '현재가', '전일비', '등락률', '거래량', '거래대금', '매수호가', '매도호가', '시가총액', 'PER', 'ROE'])

name = ''
if(select_kind == '1'):
    name = '코스피'
else:
    name = '코스닥'
date = datetime.now().strftime("%Y%m%d")
excel_filename = date + '_' + name + '_종목_' + select_theme_array[int(select_theme)-1] + '순_정렬.xlsx'



df.head(10).to_excel(excel_filename, index=False)

print(f"엑셀 파일 '{excel_filename}'이 저장되었습니다.")

file_path = 'C:\\RPAWork\\workspace\\python\\' + excel_filename
sheet_Name = 'Sheet1'

df_read = pd.read_excel(file_path, sheet_name=sheet_Name)
column_name1 = '거래량' 
column_name2 = '거래대금' 

df_read[column_name1] = df_read[column_name1].replace({',': ''}, regex=True).astype(int)
df_read[column_name2] = df_read[column_name2].replace({',': ''}, regex=True).astype(int)
column_sum = df_read[column_name1].sum()
column_sum2= df_read[column_name2].sum()

wb = load_workbook(file_path)
ws = wb[sheet_Name]

last_row = 1
while ws.cell(row=last_row, column=1).value is not None:
    last_row += 1

ws.cell(row=last_row, column=5, value=column_sum)
ws.cell(row=last_row, column=6, value=column_sum2)
wb.save(file_path)
wb.close()